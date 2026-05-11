"""
Drive → inbox ingester.

Lists audio/video, PDF, Word (.docx / .doc), plain-text (.txt), and
Excel (.xlsx / .xls) files in the configured Google Drive folder,
downloads the unprocessed ones, and routes each to the right handler:

  - Audio / video → Groq Whisper → data/requirements/inbox/meetings/*.txt
  - PDF           → pypdf text extraction → data/requirements/inbox/documents/*.txt
  - Word (.docx)  → python-docx paragraph extraction → inbox/documents/*.txt
  - Word (.doc)   → legacy OLE byte-scan → inbox/documents/*.txt
  - Plain text    → direct read → inbox/documents/*.txt
  - Excel (.xlsx) → openpyxl sheet dump → inbox/documents/*.txt
  - Excel (.xls)  → xlrd sheet dump → inbox/documents/*.txt

All outputs carry the standard 4-line header preface so the downstream
LLM splitter treats them identically to email-sourced text.

Env vars:
    GOOGLE_CREDENTIALS_PATH       Path to a service-account JSON file
                                  (Digital-Deputy-style; preferred).
                                  Resolved relative to the project root
                                  if not absolute.
    GOOGLE_SERVICE_ACCOUNT_JSON   Inline JSON key for a service account
                                  (legacy / GHA-secret style; used only
                                  if GOOGLE_CREDENTIALS_PATH is unset).
    GDRIVE_AUDIO_FOLDER_ID        The Drive folder ID (from its URL).
                                  Name kept for backward compat; the
                                  folder holds both audio and PDFs.
    GROQ_API_KEY                  Groq API key, `api` scope. Required
                                  only if the folder contains audio.
    GROQ_WHISPER_MODEL            Optional. Defaults to
                                  whisper-large-v3-turbo.

State:
    data/requirements/drive-processed.json
        { "processed": [{"file_id": ..., "name": ..., "kind": ...,
                         "ingested_at": ..., "output_path": ...}] }
    Used as the idempotency key — a Drive file ID never gets
    processed twice.

Run standalone:
    py -3 drive_ingester.py
    py -3 drive_ingester.py --dry-run
"""
from __future__ import annotations

import argparse
import io
import json
import logging
import mimetypes
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

import requests
from dotenv import load_dotenv

_HERE = Path(__file__).parent.parent  # drive/<file> -> NN root
# .env wins over inherited shell env (so a real value beats an empty
# placeholder set by the workflow). Matches agent.py's posture.
load_dotenv(_HERE / ".env", override=True)

logger = logging.getLogger(__name__)

MEETINGS_DIR = _HERE / "data" / "requirements" / "inbox" / "meetings"
DOCUMENTS_DIR = _HERE / "data" / "requirements" / "inbox" / "documents"
STATE_PATH = _HERE / "data" / "requirements" / "drive-processed.json"

# Groq's Whisper endpoint is OpenAI-compatible. Turbo is 2-3x faster
# than large-v3 with only a small quality delta; good default for
# meeting audio. Override via GROQ_WHISPER_MODEL if needed.
GROQ_URL = "https://api.groq.com/openai/v1/audio/transcriptions"
DEFAULT_MODEL = "whisper-large-v3-turbo"
GROQ_MAX_BYTES = 25 * 1024 * 1024  # 25 MB per Groq docs

AUDIO_MIME_PREFIXES = ("audio/", "video/")
AUDIO_EXT_FALLBACK = {".mp3", ".mp4", ".m4a", ".mpeg", ".mpga",
                      ".wav", ".webm", ".mkv", ".ogg", ".flac"}
PDF_MIMES = {"application/pdf"}
PDF_EXT = {".pdf"}
DOCX_MIMES = {"application/vnd.openxmlformats-officedocument.wordprocessingml.document"}
DOCX_EXT = {".docx"}
DOC_MIMES = {"application/msword"}
DOC_EXT = {".doc"}
TXT_MIMES = {"text/plain"}
TXT_EXT = {".txt"}
XLSX_MIMES = {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
XLSX_EXT = {".xlsx", ".xlsm"}
XLS_MIMES = {"application/vnd.ms-excel"}
XLS_EXT = {".xls"}


def _classify(file: dict) -> str:
    """Return 'audio', 'pdf', 'docx', 'doc', 'txt', 'xlsx', 'xls', or 'skip'."""
    mt = (file.get("mimeType") or "").lower()
    ext = Path(file.get("name") or "").suffix.lower()
    if mt.startswith(AUDIO_MIME_PREFIXES) or ext in AUDIO_EXT_FALLBACK:
        return "audio"
    if mt in PDF_MIMES or ext in PDF_EXT:
        return "pdf"
    if mt in DOC_MIMES or ext in DOC_EXT:
        return "doc"
    if mt in DOCX_MIMES or ext in DOCX_EXT:
        return "docx"
    if mt in TXT_MIMES or ext in TXT_EXT:
        return "txt"
    if mt in XLSX_MIMES or ext in XLSX_EXT:
        return "xlsx"
    if mt in XLS_MIMES or ext in XLS_EXT:
        return "xls"
    return "skip"


# ───────────────────────── state helpers ──────────────────────────

def _load_state() -> dict:
    if not STATE_PATH.exists():
        return {"processed": []}
    try:
        with open(STATE_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict) or "processed" not in data:
            return {"processed": []}
        return data
    except Exception:
        return {"processed": []}


def _save_state(state: dict) -> None:
    STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(STATE_PATH, "w", encoding="utf-8") as f:
        json.dump(state, f, indent=2, default=str)


def _already_processed(state: dict, file_id: str) -> bool:
    return any(e.get("file_id") == file_id for e in state.get("processed", []))


# ─────────────────────────── Drive ────────────────────────────────

def _load_service_account_info() -> dict:
    """Load the service-account JSON dict from either:

      1. GOOGLE_CREDENTIALS_PATH — file path (DD-style; preferred).
         Relative paths resolve against the project root.
      2. GOOGLE_SERVICE_ACCOUNT_JSON — raw JSON blob (legacy / GHA secret).

    Raises RuntimeError if neither is set or the chosen source is unparseable.
    """
    path = os.getenv("GOOGLE_CREDENTIALS_PATH")
    if path:
        p = Path(path)
        if not p.is_absolute():
            p = _HERE / p
        if not p.is_file():
            raise RuntimeError(
                f"GOOGLE_CREDENTIALS_PATH={path} does not resolve to a file "
                f"(looked at {p})"
            )
        try:
            with open(p, "r", encoding="utf-8") as f:
                return json.load(f)
        except json.JSONDecodeError as e:
            raise RuntimeError(f"{p} is not valid JSON: {e}") from e

    raw = os.getenv("GOOGLE_SERVICE_ACCOUNT_JSON")
    if not raw:
        raise RuntimeError(
            "Set GOOGLE_CREDENTIALS_PATH (preferred — points to a service-"
            "account JSON file) or GOOGLE_SERVICE_ACCOUNT_JSON (raw JSON blob)."
        )
    try:
        return json.loads(raw)
    except json.JSONDecodeError as e:
        raise RuntimeError(
            f"GOOGLE_SERVICE_ACCOUNT_JSON is not valid JSON: {e}"
        ) from e


def _drive_service():
    """Build a Drive v3 client from the JSON service-account key."""
    info = _load_service_account_info()
    from google.oauth2 import service_account  # lazy import
    from googleapiclient.discovery import build
    creds = service_account.Credentials.from_service_account_info(
        info, scopes=["https://www.googleapis.com/auth/drive.readonly",
                      "https://www.googleapis.com/auth/drive"])
    return build("drive", "v3", credentials=creds, cache_discovery=False)


def _list_ingestable_files(drive, folder_id: str) -> list[dict]:
    """Return non-trashed children we can handle (audio/video, PDF, docx, txt)."""
    q = f"'{folder_id}' in parents and trashed = false"
    files: list[dict] = []
    page_token = None
    while True:
        resp = drive.files().list(
            q=q,
            fields=("nextPageToken,"
                    " files(id, name, mimeType, size, createdTime)"),
            pageSize=100,
            pageToken=page_token,
            supportsAllDrives=True,
            includeItemsFromAllDrives=True,
        ).execute()
        for f in resp.get("files", []):
            if _classify(f) != "skip":
                files.append(f)
        page_token = resp.get("nextPageToken")
        if not page_token:
            break
    return files


def _download_file(drive, file_id: str, out_path: Path) -> None:
    """Stream-download a Drive file to disk."""
    from googleapiclient.http import MediaIoBaseDownload  # lazy
    request = drive.files().get_media(fileId=file_id, supportsAllDrives=True)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "wb") as fh:
        downloader = MediaIoBaseDownload(fh, request, chunksize=1024 * 1024)
        done = False
        while not done:
            _status, done = downloader.next_chunk()


# ─────────────────────────── Groq ─────────────────────────────────

def _transcribe_via_groq(audio_path: Path) -> str:
    """POST the audio file to Groq's Whisper endpoint. Returns text."""
    api_key = os.getenv("GROQ_API_KEY")
    if not api_key:
        raise RuntimeError("GROQ_API_KEY not set")
    model = os.getenv("GROQ_WHISPER_MODEL") or DEFAULT_MODEL

    size = audio_path.stat().st_size
    if size > GROQ_MAX_BYTES:
        raise RuntimeError(
            f"{audio_path.name} is {size / 1e6:.1f} MB — exceeds Groq's "
            f"{GROQ_MAX_BYTES / 1e6:.0f} MB limit. Shorten the recording "
            f"or split the file before uploading to Drive."
        )

    mime = mimetypes.guess_type(str(audio_path))[0] or "application/octet-stream"
    with open(audio_path, "rb") as f:
        files = {"file": (audio_path.name, f, mime)}
        data = {
            "model": model,
            # Leaving language empty lets Whisper auto-detect.
            # Whisper's Bangla recognition is strong, and the downstream
            # LLM splitter already forces English output.
            "response_format": "text",
        }
        r = requests.post(
            GROQ_URL,
            headers={"Authorization": f"Bearer {api_key}"},
            files=files, data=data,
            timeout=600,  # 10 min — a long meeting can take a while
        )
    if not r.ok:
        raise RuntimeError(f"Groq transcription failed {r.status_code}: "
                           f"{r.text[:400]}")
    # response_format=text returns raw text body, not JSON.
    return r.text.strip()


# ─────────────────────────── PDF ──────────────────────────────────

def _extract_pdf_text(pdf_path: Path) -> str:
    """Pure-Python text extraction via pypdf. Tolerates most standard
    (non-scanned) PDFs. Scanned / image-only PDFs yield empty output
    and would need OCR (not wired here)."""
    from pypdf import PdfReader  # lazy import
    reader = PdfReader(str(pdf_path))
    parts: list[str] = []
    for i, page in enumerate(reader.pages, 1):
        try:
            parts.append(page.extract_text() or "")
        except Exception as e:
            parts.append(f"[page {i} extraction failed: {e}]")
    text = "\n\n".join(parts).strip()
    return text


# ─────────────────────────── Word / text ──────────────────────────

def _extract_docx_text(docx_path: Path) -> str:
    """Extract paragraph text from a Word document via python-docx."""
    from docx import Document  # lazy import — python-docx
    doc = Document(str(docx_path))
    parts = [para.text for para in doc.paragraphs if para.text.strip()]
    return "\n\n".join(parts)


def _extract_doc_text(doc_path: Path) -> str:
    """Best-effort text extraction from legacy .doc (OLE compound).

    Pure Python options for .doc are limited. This walks the binary and
    pulls out printable ASCII / UTF-8 runs ≥ 4 chars, which catches the
    bulk of plaintext content in most .doc files. Result is rougher
    than .docx extraction — for a clean output, the user should re-save
    the file as .docx.
    """
    raw = doc_path.read_bytes()
    out: list[str] = []
    buf: list[int] = []
    for b in raw:
        if 32 <= b < 127 or b in (9, 10, 13):
            buf.append(b)
        else:
            if len(buf) >= 4:
                out.append(bytes(buf).decode("utf-8", "replace"))
            buf = []
    if len(buf) >= 4:
        out.append(bytes(buf).decode("utf-8", "replace"))
    text = "\n".join(s for s in out if s.strip())
    if not text:
        return ("[legacy .doc — no plain text found via byte scan; "
                "please re-save as .docx for a clean extract]")
    return text


def _extract_txt_text(txt_path: Path) -> str:
    return txt_path.read_text(encoding="utf-8", errors="replace")


# ─────────────────────────── Excel ────────────────────────────────

def _extract_xlsx_text(xlsx_path: Path) -> str:
    """Plain-text dump of every sheet via openpyxl. Empty cells -> ''."""
    from openpyxl import load_workbook  # lazy
    wb = load_workbook(str(xlsx_path), data_only=True, read_only=True)
    parts: list[str] = []
    for ws in wb.worksheets:
        parts.append(f"=== Sheet: {ws.title} ===")
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in row]
            if any(c.strip() for c in cells):
                parts.append(" | ".join(cells))
        parts.append("")
    return "\n".join(parts).strip()


def _extract_xls_text(xls_path: Path) -> str:
    """Plain-text dump of every sheet via xlrd (legacy .xls)."""
    import xlrd  # lazy
    book = xlrd.open_workbook(str(xls_path))
    parts: list[str] = []
    for sheet in book.sheets():
        parts.append(f"=== Sheet: {sheet.name} ===")
        for r in range(sheet.nrows):
            cells = [str(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
            if any(c.strip() for c in cells):
                parts.append(" | ".join(cells))
        parts.append("")
    return "\n".join(parts).strip()


# ─────────────────────────── Main ─────────────────────────────────

def _preface(file: dict, source_label: str) -> str:
    return (
        f"# source: {source_label}\n"
        f"# from: drive:{file.get('id')}\n"
        f"# received: {file.get('createdTime') or datetime.now(timezone.utc).isoformat()}\n"
        f"# subject: {file.get('name')}\n\n"
    )


def _safe_filename(original: str, default_stem: str = "file") -> str:
    stem = Path(original).stem
    safe = "".join(c if c.isalnum() or c in "-_" else "-" for c in stem).strip("-")
    date = datetime.now().strftime("%Y-%m-%d")
    return f"{date}-{safe or default_stem}.txt"


def process_new_drive_files(dry_run: bool = False) -> dict:
    folder_id = os.getenv("GDRIVE_AUDIO_FOLDER_ID")
    if not folder_id:
        return {"error": "GDRIVE_AUDIO_FOLDER_ID not set",
                "processed": 0, "files": []}

    drive = _drive_service()
    listed = _list_ingestable_files(drive, folder_id)
    logger.info(f"Drive folder {folder_id}: found {len(listed)} "
                f"ingestable file(s) (audio + PDF + docx + doc + txt + xlsx + xls)")

    state = _load_state()
    pending = [f for f in listed if not _already_processed(state, f["id"])]
    if not pending:
        return {"processed": 0, "files": [], "skipped_already_done": len(listed)}

    MEETINGS_DIR.mkdir(parents=True, exist_ok=True)
    DOCUMENTS_DIR.mkdir(parents=True, exist_ok=True)
    tmp_dir = _HERE / "data" / "requirements" / "_drive_tmp"
    tmp_dir.mkdir(parents=True, exist_ok=True)

    results: list[dict] = []
    for f in pending:
        fid = f["id"]
        name = f["name"]
        kind = _classify(f)
        tmp_path = tmp_dir / f"{fid}-{Path(name).name}"

        logger.info(f"[{kind}] downloading {name} (id={fid}, "
                    f"{int(f.get('size') or 0)/1e6:.1f} MB)")
        try:
            _download_file(drive, fid, tmp_path)
        except Exception as e:
            logger.exception(f"download failed for {name}")
            results.append({"file": name, "id": fid, "kind": kind,
                            "error": f"download: {e}"})
            continue

        if dry_run:
            logger.info(f"[dry-run] would process {name} as {kind}")
            results.append({"file": name, "id": fid, "kind": kind,
                            "dry_run": True})
            try: tmp_path.unlink()
            except Exception: pass
            continue

        # Extract content per kind.
        text = ""
        if kind == "audio":
            try:
                logger.info(f"transcribing {name} via Groq Whisper...")
                text = _transcribe_via_groq(tmp_path)
            except Exception as e:
                logger.exception(f"transcription failed for {name}")
                results.append({"file": name, "id": fid, "kind": kind,
                                "error": f"groq: {e}"})
                try: tmp_path.unlink()
                except Exception: pass
                continue
        elif kind == "pdf":
            try:
                logger.info(f"extracting text from {name} via pypdf...")
                text = _extract_pdf_text(tmp_path)
                if not text:
                    logger.warning(f"pypdf returned empty text for {name} — "
                                   f"likely a scanned PDF; OCR not wired")
            except Exception as e:
                logger.exception(f"pdf extraction failed for {name}")
                results.append({"file": name, "id": fid, "kind": kind,
                                "error": f"pypdf: {e}"})
                try: tmp_path.unlink()
                except Exception: pass
                continue
        elif kind == "docx":
            try:
                logger.info(f"extracting text from {name} via python-docx...")
                text = _extract_docx_text(tmp_path)
                if not text:
                    logger.warning(f"python-docx returned empty text for {name}")
            except Exception as e:
                logger.exception(f"docx extraction failed for {name}")
                results.append({"file": name, "id": fid, "kind": kind,
                                "error": f"docx: {e}"})
                try: tmp_path.unlink()
                except Exception: pass
                continue
        elif kind == "txt":
            try:
                logger.info(f"reading plain text {name}...")
                text = _extract_txt_text(tmp_path)
            except Exception as e:
                logger.exception(f"txt read failed for {name}")
                results.append({"file": name, "id": fid, "kind": kind,
                                "error": f"txt: {e}"})
                try: tmp_path.unlink()
                except Exception: pass
                continue
        elif kind == "doc":
            try:
                logger.info(f"extracting text from {name} (legacy .doc, byte scan)...")
                text = _extract_doc_text(tmp_path)
            except Exception as e:
                logger.exception(f"doc extraction failed for {name}")
                results.append({"file": name, "id": fid, "kind": kind,
                                "error": f"doc: {e}"})
                try: tmp_path.unlink()
                except Exception: pass
                continue
        elif kind == "xlsx":
            try:
                logger.info(f"extracting text from {name} via openpyxl...")
                text = _extract_xlsx_text(tmp_path)
                if not text:
                    logger.warning(f"openpyxl returned empty text for {name}")
            except Exception as e:
                logger.exception(f"xlsx extraction failed for {name}")
                results.append({"file": name, "id": fid, "kind": kind,
                                "error": f"xlsx: {e}"})
                try: tmp_path.unlink()
                except Exception: pass
                continue
        elif kind == "xls":
            try:
                logger.info(f"extracting text from {name} via xlrd...")
                text = _extract_xls_text(tmp_path)
                if not text:
                    logger.warning(f"xlrd returned empty text for {name}")
            except Exception as e:
                logger.exception(f"xls extraction failed for {name}")
                results.append({"file": name, "id": fid, "kind": kind,
                                "error": f"xls: {e}"})
                try: tmp_path.unlink()
                except Exception: pass
                continue
        else:
            # Shouldn't happen — listed files already filtered.
            try: tmp_path.unlink()
            except Exception: pass
            continue

        # Route to the right inbox subfolder.
        if kind == "audio":
            source_label = "meetings"
            out_dir = MEETINGS_DIR
            default_stem = "audio"
        else:  # pdf, docx, txt
            source_label = "documents"
            out_dir = DOCUMENTS_DIR
            default_stem = "document"

        out_name = _safe_filename(name, default_stem=default_stem)
        out_path = out_dir / out_name
        if out_path.exists():
            out_path = out_dir / f"{out_path.stem}-{fid[:6]}.txt"
        out_path.write_text(_preface(f, source_label) + text, encoding="utf-8")
        logger.info(f"wrote {kind} -> {out_path.relative_to(_HERE).as_posix()}")

        state["processed"].append({
            "file_id": fid,
            "name": name,
            "kind": kind,
            "ingested_at": datetime.now(timezone.utc).isoformat(),
            "output_path": str(out_path.relative_to(_HERE).as_posix()),
        })
        _save_state(state)

        try: tmp_path.unlink()
        except Exception: pass

        results.append({
            "file": name,
            "id": fid,
            "kind": kind,
            "output": str(out_path.relative_to(_HERE).as_posix()),
            "chars": len(text),
        })

    return {
        "processed": sum(1 for r in results if "output" in r),
        "errors": sum(1 for r in results if "error" in r),
        "files": results,
    }


def _cli() -> None:
    logging.basicConfig(level=logging.INFO,
                        format="%(asctime)s [%(levelname)s] %(message)s")
    p = argparse.ArgumentParser()
    p.add_argument("--dry-run", action="store_true",
                   help="List + download pending files but do not run "
                        "Whisper / pypdf.")
    args = p.parse_args()
    out = process_new_drive_files(dry_run=args.dry_run)
    print(json.dumps(out, indent=2, default=str))
    if out.get("error"):
        sys.exit(1)


if __name__ == "__main__":
    _cli()
